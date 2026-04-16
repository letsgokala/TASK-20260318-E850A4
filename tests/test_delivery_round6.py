"""Round-6 regression tests — semantic correctness.

Covers:
1. correction_rate: resolved corrections no longer inflate the metric.
2. compliance report: Materials Complete uses required-checklist logic.
3. quality validation: optional missing items are warnings, not failures.
4. batch review: unsupported actions are rejected by the schema.
5. JWT base64url parsing (frontend, vitest companion).
"""
from __future__ import annotations

import os
import uuid
from datetime import datetime, timedelta, timezone
from decimal import Decimal

import pytest
from httpx import AsyncClient
from sqlalchemy import select

from app.models.checklist_item import ChecklistItem
from app.models.collection_batch import CollectionBatch
from app.models.material import Material, MaterialVersion, MaterialVersionStatus
from app.models.registration import Registration, RegistrationStatus
from tests.conftest import make_token


# ── Helpers ────────────────────────────────────────────────────────────────

async def _batch(db_session, admin_user):
    b = CollectionBatch(
        name="R6 Batch",
        submission_deadline=datetime.now(timezone.utc) + timedelta(days=30),
        created_by=admin_user.id,
    )
    db_session.add(b)
    await db_session.commit()
    await db_session.refresh(b)
    return b


async def _reg(db_session, applicant_user, batch, status=RegistrationStatus.SUBMITTED):
    r = Registration(
        batch_id=batch.id,
        applicant_id=applicant_user.id,
        status=status,
        title="R6",
        activity_type="research",
        description="d",
        applicant_name="A",
    )
    db_session.add(r)
    await db_session.commit()
    await db_session.refresh(r)
    return r


# ═══════════════════════════════════════════════════════════════════════════
# 1. Correction-rate: resolved corrections should NOT inflate the metric
# ═══════════════════════════════════════════════════════════════════════════

@pytest.mark.asyncio
async def test_correction_rate_excludes_resolved_corrections(
    client: AsyncClient, reviewer_user, applicant_user, admin_user, db_session,
):
    """If a material's v1 is needs_correction but v2 is submitted, the
    registration must NOT count toward correction_rate."""
    batch = await _batch(db_session, admin_user)
    ci = ChecklistItem(
        batch_id=batch.id, label="Evidence", is_required=True, sort_order=1,
    )
    db_session.add(ci)
    await db_session.flush()

    reg = await _reg(db_session, applicant_user, batch)
    mat = Material(registration_id=reg.id, checklist_item_id=ci.id)
    db_session.add(mat)
    await db_session.flush()

    # v1: needs_correction (historical — resolved by v2).
    v1 = MaterialVersion(
        material_id=mat.id,
        version_number=1,
        original_filename="old.pdf",
        mime_type="application/pdf",
        file_size_bytes=100,
        sha256_hash="a" * 64,
        storage_path="/tmp/old.pdf",
        status=MaterialVersionStatus.NEEDS_CORRECTION,
        uploaded_by=applicant_user.id,
    )
    # v2: submitted (corrected — resolves the correction).
    v2 = MaterialVersion(
        material_id=mat.id,
        version_number=2,
        original_filename="new.pdf",
        mime_type="application/pdf",
        file_size_bytes=200,
        sha256_hash="b" * 64,
        storage_path="/tmp/new.pdf",
        status=MaterialVersionStatus.SUBMITTED,
        uploaded_by=applicant_user.id,
    )
    db_session.add_all([v1, v2])
    await db_session.commit()

    resp = await client.get("/api/v1/metrics", headers=make_token(reviewer_user))
    assert resp.status_code == 200
    data = resp.json()
    # The only registration's correction is resolved → rate must be 0.
    assert float(data["correction_rate"]["value"]) == 0.0


@pytest.mark.asyncio
async def test_correction_rate_counts_unresolved(
    client: AsyncClient, reviewer_user, applicant_user, admin_user, db_session,
):
    """A material with only a v1=needs_correction (no newer version)
    IS unresolved and must count toward the rate."""
    batch = await _batch(db_session, admin_user)
    ci = ChecklistItem(
        batch_id=batch.id, label="Evidence", is_required=True, sort_order=1,
    )
    db_session.add(ci)
    await db_session.flush()

    reg = await _reg(db_session, applicant_user, batch)
    mat = Material(registration_id=reg.id, checklist_item_id=ci.id)
    db_session.add(mat)
    await db_session.flush()

    v1 = MaterialVersion(
        material_id=mat.id,
        version_number=1,
        original_filename="bad.pdf",
        mime_type="application/pdf",
        file_size_bytes=100,
        sha256_hash="c" * 64,
        storage_path="/tmp/bad.pdf",
        status=MaterialVersionStatus.NEEDS_CORRECTION,
        uploaded_by=applicant_user.id,
    )
    db_session.add(v1)
    await db_session.commit()

    resp = await client.get("/api/v1/metrics", headers=make_token(reviewer_user))
    assert resp.status_code == 200
    data = resp.json()
    # 1 registration, 1 unresolved correction → 100%
    assert float(data["correction_rate"]["value"]) == 100.0


# ═══════════════════════════════════════════════════════════════════════════
# 2. Compliance report: required-checklist completeness
# ═══════════════════════════════════════════════════════════════════════════

@pytest.mark.asyncio
async def test_compliance_report_required_item_missing_is_incomplete(
    db_session, admin_user, applicant_user,
):
    """If a required checklist item has no material version uploaded,
    the compliance report must say 'No' for Materials Complete."""
    from app.reports.generator import generate_compliance_report
    from openpyxl import load_workbook

    batch = await _batch(db_session, admin_user)
    ci_req = ChecklistItem(
        batch_id=batch.id, label="Required Doc", is_required=True, sort_order=1,
    )
    ci_opt = ChecklistItem(
        batch_id=batch.id, label="Optional Doc", is_required=False, sort_order=2,
    )
    db_session.add_all([ci_req, ci_opt])
    await db_session.flush()

    # Registration with ONLY the optional item uploaded.
    reg = await _reg(db_session, applicant_user, batch)
    mat_opt = Material(registration_id=reg.id, checklist_item_id=ci_opt.id)
    db_session.add(mat_opt)
    await db_session.flush()
    ver_opt = MaterialVersion(
        material_id=mat_opt.id,
        version_number=1,
        original_filename="opt.pdf",
        mime_type="application/pdf",
        file_size_bytes=100,
        sha256_hash="d" * 64,
        storage_path="/tmp/opt.pdf",
        status=MaterialVersionStatus.SUBMITTED,
        uploaded_by=applicant_user.id,
    )
    db_session.add(ver_opt)
    await db_session.commit()

    path = await generate_compliance_report(db_session, batch.id, None, None)
    try:
        wb = load_workbook(path)
        ws = wb["Compliance"]
        # Row 2 should be our registration.
        materials_complete = ws.cell(row=2, column=4).value
        assert materials_complete == "No", (
            "Required item missing → Materials Complete must be 'No'"
        )
    finally:
        if os.path.isfile(path):
            os.remove(path)


@pytest.mark.asyncio
async def test_compliance_report_all_required_present_is_complete(
    db_session, admin_user, applicant_user,
):
    """When all required items have valid uploads, Materials Complete = Yes
    regardless of missing optional items."""
    from app.reports.generator import generate_compliance_report
    from openpyxl import load_workbook

    batch = await _batch(db_session, admin_user)
    ci_req = ChecklistItem(
        batch_id=batch.id, label="Required Doc", is_required=True, sort_order=1,
    )
    ci_opt = ChecklistItem(
        batch_id=batch.id, label="Optional Doc", is_required=False, sort_order=2,
    )
    db_session.add_all([ci_req, ci_opt])
    await db_session.flush()

    reg = await _reg(db_session, applicant_user, batch)
    mat_req = Material(registration_id=reg.id, checklist_item_id=ci_req.id)
    db_session.add(mat_req)
    await db_session.flush()
    ver_req = MaterialVersion(
        material_id=mat_req.id,
        version_number=1,
        original_filename="req.pdf",
        mime_type="application/pdf",
        file_size_bytes=100,
        sha256_hash="e" * 64,
        storage_path="/tmp/req.pdf",
        status=MaterialVersionStatus.SUBMITTED,
        uploaded_by=applicant_user.id,
    )
    db_session.add(ver_req)
    await db_session.commit()

    path = await generate_compliance_report(db_session, batch.id, None, None)
    try:
        wb = load_workbook(path)
        ws = wb["Compliance"]
        materials_complete = ws.cell(row=2, column=4).value
        assert materials_complete == "Yes", (
            "All required items present → Materials Complete must be 'Yes' "
            "even when optional items are missing"
        )
    finally:
        if os.path.isfile(path):
            os.remove(path)


# ═══════════════════════════════════════════════════════════════════════════
# 3. Quality validation: optional items are warnings, not failures
# ═══════════════════════════════════════════════════════════════════════════

@pytest.mark.asyncio
async def test_validation_optional_missing_is_warning_not_failure(
    db_session, admin_user, applicant_user,
):
    """A missing optional checklist item must produce a WARNING result,
    not a FAIL. The summary's ``failed`` count must not include it."""
    from app.api.v1.quality_validation import _run_all_rules
    from app.models.quality_validation import ValidationStatus

    batch = await _batch(db_session, admin_user)
    ci_req = ChecklistItem(
        batch_id=batch.id, label="Required", is_required=True, sort_order=1,
    )
    ci_opt = ChecklistItem(
        batch_id=batch.id, label="Optional", is_required=False, sort_order=2,
    )
    db_session.add_all([ci_req, ci_opt])
    await db_session.flush()

    reg = await _reg(db_session, applicant_user, batch)

    # Upload material for the required item only.
    mat_req = Material(registration_id=reg.id, checklist_item_id=ci_req.id)
    db_session.add(mat_req)
    await db_session.flush()
    ver = MaterialVersion(
        material_id=mat_req.id,
        version_number=1,
        original_filename="r.pdf",
        mime_type="application/pdf",
        file_size_bytes=100,
        sha256_hash="f" * 64,
        storage_path="/tmp/r.pdf",
        status=MaterialVersionStatus.SUBMITTED,
        uploaded_by=applicant_user.id,
    )
    db_session.add(ver)
    await db_session.commit()

    from app.models.user import User
    results = await _run_all_rules(reg, applicant_user, db_session)

    # Find the material_complete result for the optional item.
    opt_result = next(
        r for r in results
        if "material_complete" in r.rule_name and str(ci_opt.id) in r.rule_name
    )
    assert opt_result.status == ValidationStatus.WARNING, (
        f"Optional missing item must be WARNING, got {opt_result.status}"
    )

    # The required item should be a PASS since it has a valid upload.
    req_result = next(
        r for r in results
        if "material_complete" in r.rule_name and str(ci_req.id) in r.rule_name
    )
    assert req_result.status == ValidationStatus.PASS

    # Summary: the optional item must not count in the "failed" bucket.
    failed = sum(1 for r in results if r.status == ValidationStatus.FAIL)
    material_fails = [
        r for r in results
        if "material_complete" in r.rule_name and r.status == ValidationStatus.FAIL
    ]
    assert len(material_fails) == 0, (
        "No material_complete rule should FAIL when only optional items are missing"
    )


# ═══════════════════════════════════════════════════════════════════════════
# 4. Batch review: unsupported actions rejected
# ═══════════════════════════════════════════════════════════════════════════

@pytest.mark.asyncio
async def test_batch_review_rejects_canceled_action(
    client: AsyncClient, reviewer_user, applicant_user, admin_user, db_session,
):
    """Batch review must NOT accept 'canceled' — only
    approved/rejected/waitlisted per the API spec."""
    batch = await _batch(db_session, admin_user)
    reg = await _reg(db_session, applicant_user, batch)

    resp = await client.post(
        "/api/v1/reviews/batch",
        json={
            "registration_ids": [str(reg.id)],
            "action": "canceled",
        },
        headers=make_token(reviewer_user),
    )
    assert resp.status_code == 422, (
        "batch review with action='canceled' must be rejected by schema validation"
    )


@pytest.mark.asyncio
async def test_batch_review_rejects_draft_action(
    client: AsyncClient, reviewer_user, applicant_user, admin_user, db_session,
):
    batch = await _batch(db_session, admin_user)
    reg = await _reg(db_session, applicant_user, batch)

    resp = await client.post(
        "/api/v1/reviews/batch",
        json={
            "registration_ids": [str(reg.id)],
            "action": "draft",
        },
        headers=make_token(reviewer_user),
    )
    assert resp.status_code == 422


@pytest.mark.asyncio
async def test_batch_review_accepts_approved(
    client: AsyncClient, reviewer_user, applicant_user, admin_user, db_session,
):
    """Positive case: 'approved' is a valid batch action and should succeed."""
    batch = await _batch(db_session, admin_user)
    reg = await _reg(db_session, applicant_user, batch)

    resp = await client.post(
        "/api/v1/reviews/batch",
        json={
            "registration_ids": [str(reg.id)],
            "action": "approved",
        },
        headers=make_token(reviewer_user),
    )
    assert resp.status_code == 200


# ═══════════════════════════════════════════════════════════════════════════
# 5. Static structural guards
# ═══════════════════════════════════════════════════════════════════════════

def test_correction_rate_uses_latest_version_logic():
    """The correction-rate query must use the latest-version subquery
    so resolved corrections are excluded."""
    import inspect
    from app.api.v1.metrics import _unresolved_correction_count_query
    src = inspect.getsource(_unresolved_correction_count_query)
    assert "func.max(MaterialVersion.version_number)" in src
    assert "NEEDS_CORRECTION" in src


def test_batch_review_schema_restricts_actions():
    """BatchReviewAction must only accept approved/rejected/waitlisted."""
    from app.schemas.review import BatchReviewAction
    assert set(BatchReviewAction) == {
        BatchReviewAction.APPROVED,
        BatchReviewAction.REJECTED,
        BatchReviewAction.WAITLISTED,
    }


def test_jwt_parsing_is_base64url_safe():
    """LoginPage.vue must normalise base64url before calling atob."""
    import pathlib
    src = (
        pathlib.Path(__file__).resolve().parent.parent
        / "frontend" / "src" / "views" / "LoginPage.vue"
    ).read_text()
    assert "replace(/-/g" in src, "JWT parsing must handle base64url '-' chars"
    assert "replace(/_/g" in src, "JWT parsing must handle base64url '_' chars"
    # The old brittle pattern should be gone.
    assert "atob(token.split" not in src, (
        "The direct atob(token.split('.')[1]) pattern should be replaced "
        "by the base64url-normalised version"
    )
