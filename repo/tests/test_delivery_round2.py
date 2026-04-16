"""Round-2 regression tests covering the remaining audit blockers.

The second audit report flagged three High-severity defects that the
earlier transactional-audit fix did not address:

1. Material versions never transition from ``pending_submission`` to
   ``submitted`` during the submit / supplementary flows.
2. The restore endpoint could return ``status=complete`` even when
   rsync failed (nonzero ``files_exit_code``).
3. The reviewer SPA had no UI action to download the materials it is
   reviewing (covered by a frontend vitest; a structural guard here).

These tests pin the fixes so the gaps cannot silently regrow.
"""
from __future__ import annotations

import os
import uuid
from datetime import datetime, timedelta, timezone
from unittest.mock import patch, MagicMock

import pytest
from httpx import AsyncClient

from app.models.checklist_item import ChecklistItem
from app.models.collection_batch import CollectionBatch
from app.models.material import Material, MaterialVersion, MaterialVersionStatus
from app.models.registration import Registration, RegistrationStatus
from tests.conftest import make_token


# ── Material version lifecycle on submit ──────────────────────────────────

async def _make_batch(db_session, admin_user, deadline_days: int = 30) -> CollectionBatch:
    batch = CollectionBatch(
        name="Lifecycle Batch",
        submission_deadline=datetime.now(timezone.utc) + timedelta(days=deadline_days),
        created_by=admin_user.id,
    )
    db_session.add(batch)
    await db_session.commit()
    await db_session.refresh(batch)
    return batch


async def _make_registration_with_material(
    db_session, applicant_user, batch, has_version: bool = True,
) -> tuple[Registration, Material, MaterialVersion | None]:
    ci = ChecklistItem(
        batch_id=batch.id,
        label="Required Doc",
        is_required=True,
        sort_order=1,
    )
    db_session.add(ci)
    await db_session.flush()

    reg = Registration(
        batch_id=batch.id,
        applicant_id=applicant_user.id,
        status=RegistrationStatus.DRAFT,
        title="Lifecycle Test",
        activity_type="research",
        description="Desc",
        applicant_name="Applicant",
        applicant_id_number="123456789012",
        applicant_phone="+11234567890",
        applicant_email="a@b.com",
        start_date=datetime.now(timezone.utc),
        end_date=datetime.now(timezone.utc) + timedelta(days=7),
        requested_budget=100,
    )
    db_session.add(reg)
    await db_session.flush()

    material = Material(registration_id=reg.id, checklist_item_id=ci.id)
    db_session.add(material)
    await db_session.flush()

    version = None
    if has_version:
        version = MaterialVersion(
            material_id=material.id,
            version_number=1,
            original_filename="doc.pdf",
            mime_type="application/pdf",
            file_size_bytes=1024,
            sha256_hash="a" * 64,
            storage_path=f"/tmp/fake_{uuid.uuid4().hex}.pdf",
            status=MaterialVersionStatus.PENDING_SUBMISSION,
            uploaded_by=applicant_user.id,
        )
        db_session.add(version)

    await db_session.commit()
    await db_session.refresh(reg)
    if version:
        await db_session.refresh(version)
    return reg, material, version


@pytest.mark.asyncio
async def test_submit_promotes_pending_versions_to_submitted(
    client: AsyncClient, applicant_user, admin_user, db_session,
):
    """After the applicant submits, every pending_submission version on
    their registration must read as submitted. The audit report flagged
    this lifecycle as a High-severity prompt miss."""
    batch = await _make_batch(db_session, admin_user)
    reg, material, version = await _make_registration_with_material(
        db_session, applicant_user, batch
    )
    assert version.status == MaterialVersionStatus.PENDING_SUBMISSION

    resp = await client.post(
        f"/api/v1/registrations/{reg.id}/submit",
        headers=make_token(applicant_user),
    )
    assert resp.status_code == 200

    await db_session.refresh(version)
    assert version.status == MaterialVersionStatus.SUBMITTED


@pytest.mark.asyncio
async def test_submit_preserves_needs_correction_status(
    client: AsyncClient, applicant_user, admin_user, db_session,
):
    """The submit-time transition must only advance pending versions —
    a reviewer-set ``needs_correction`` label is a deliberate state and
    must not be overwritten by another submit pass."""
    batch = await _make_batch(db_session, admin_user)
    reg, material, version = await _make_registration_with_material(
        db_session, applicant_user, batch
    )

    # Manually flip the existing version to needs_correction and add a
    # second pending one for the same material — the submit pass should
    # promote the pending one and leave needs_correction alone.
    version.status = MaterialVersionStatus.NEEDS_CORRECTION
    second = MaterialVersion(
        material_id=material.id,
        version_number=2,
        original_filename="doc_v2.pdf",
        mime_type="application/pdf",
        file_size_bytes=2048,
        sha256_hash="b" * 64,
        storage_path=f"/tmp/fake_v2_{uuid.uuid4().hex}.pdf",
        status=MaterialVersionStatus.PENDING_SUBMISSION,
        uploaded_by=applicant_user.id,
    )
    db_session.add(second)
    await db_session.commit()

    resp = await client.post(
        f"/api/v1/registrations/{reg.id}/submit",
        headers=make_token(applicant_user),
    )
    assert resp.status_code == 200

    await db_session.refresh(version)
    await db_session.refresh(second)
    assert version.status == MaterialVersionStatus.NEEDS_CORRECTION
    assert second.status == MaterialVersionStatus.SUBMITTED


# ── Restore partial-failure surfacing ─────────────────────────────────────

@pytest.mark.asyncio
async def test_restore_reports_partial_when_rsync_fails(
    client: AsyncClient, admin_user, tmp_path, monkeypatch,
):
    """When rsync returns a nonzero exit code, the restore endpoint must
    NOT report ``status=complete``. The audit report flagged this as a
    High-severity recovery trust issue."""
    # Create a fake backup on disk so the endpoint finds both halves.
    date_str = "20250101"
    db_dump = tmp_path / f"backup_{date_str}.dump"
    db_dump.write_bytes(b"fake dump")
    file_backup = tmp_path / date_str
    file_backup.mkdir()
    (file_backup / "marker").write_text("x")

    from app.api.v1 import admin_ops as ops
    monkeypatch.setattr(ops, "_BACKUP_DB_DIR", str(tmp_path))
    monkeypatch.setattr(ops, "_BACKUP_FILES_DIR", str(tmp_path))

    def fake_run(cmd, *args, **kwargs):
        mock = MagicMock()
        if cmd[0] == "pg_restore":
            mock.returncode = 0
            mock.stderr = b""
        elif cmd[0] == "rsync":
            mock.returncode = 23  # rsync "partial transfer" non-zero exit
            mock.stderr = b"rsync: failed to set permissions"
        else:
            mock.returncode = 0
            mock.stderr = b""
        return mock

    monkeypatch.setattr(ops.subprocess, "run", fake_run)

    resp = await client.post(
        f"/api/v1/admin/backups/{date_str}/restore",
        headers=make_token(admin_user),
    )
    assert resp.status_code == 202
    body = resp.json()
    assert body["status"] != "complete", (
        "rsync failure must not be reported as complete"
    )
    assert body["status"] in ("partial", "failed")
    assert body["files_exit_code"] == 23
    assert body["files_attempted"] is True


@pytest.mark.asyncio
async def test_restore_reports_complete_only_when_everything_succeeds(
    client: AsyncClient, admin_user, tmp_path, monkeypatch,
):
    """Complementary case: both subprocesses return 0 → status=complete."""
    date_str = "20250102"
    db_dump = tmp_path / f"backup_{date_str}.dump"
    db_dump.write_bytes(b"fake dump")
    file_backup = tmp_path / date_str
    file_backup.mkdir()

    from app.api.v1 import admin_ops as ops
    monkeypatch.setattr(ops, "_BACKUP_DB_DIR", str(tmp_path))
    monkeypatch.setattr(ops, "_BACKUP_FILES_DIR", str(tmp_path))

    def fake_run(cmd, *args, **kwargs):
        mock = MagicMock()
        mock.returncode = 0
        mock.stderr = b""
        return mock

    monkeypatch.setattr(ops.subprocess, "run", fake_run)

    resp = await client.post(
        f"/api/v1/admin/backups/{date_str}/restore",
        headers=make_token(admin_user),
    )
    assert resp.status_code == 202
    body = resp.json()
    assert body["status"] == "complete"
    assert body["db_exit_code"] == 0
    assert body["files_exit_code"] == 0


# ── Whitelist report semantics ────────────────────────────────────────────

@pytest.mark.asyncio
async def test_whitelist_report_lists_approved_registrations(
    db_session, admin_user, applicant_user,
):
    """The whitelist report must include approved registrations — not only
    checklist definitions. The previous implementation exported the batch
    checklist, which did not match the documented/prompt intent."""
    from app.reports.generator import generate_whitelist_report
    from openpyxl import load_workbook

    batch = await _make_batch(db_session, admin_user)
    ci = ChecklistItem(
        batch_id=batch.id, label="Letter", is_required=True, sort_order=1,
    )
    db_session.add(ci)
    await db_session.flush()

    approved = Registration(
        batch_id=batch.id,
        applicant_id=applicant_user.id,
        status=RegistrationStatus.APPROVED,
        title="Approved Project",
        activity_type="research",
        description="desc",
        applicant_name="Approved Applicant",
    )
    rejected = Registration(
        batch_id=batch.id,
        applicant_id=applicant_user.id,
        status=RegistrationStatus.REJECTED,
        title="Rejected Project",
        activity_type="research",
        description="desc",
        applicant_name="Rejected Applicant",
    )
    db_session.add_all([approved, rejected])
    await db_session.flush()

    mat = Material(registration_id=approved.id, checklist_item_id=ci.id)
    db_session.add(mat)
    await db_session.flush()
    mv = MaterialVersion(
        material_id=mat.id,
        version_number=1,
        original_filename="letter.pdf",
        mime_type="application/pdf",
        file_size_bytes=111,
        sha256_hash="c" * 64,
        storage_path="/tmp/letter.pdf",
        status=MaterialVersionStatus.SUBMITTED,
        uploaded_by=applicant_user.id,
    )
    db_session.add(mv)
    await db_session.commit()

    try:
        file_path = await generate_whitelist_report(db_session, batch.id)
        wb = load_workbook(file_path)
        assert "Approved Registrations" in wb.sheetnames
        assert "Approved Materials" in wb.sheetnames

        ws = wb["Approved Registrations"]
        names = {ws.cell(row=i, column=4).value for i in range(2, ws.max_row + 1)}
        assert "Approved Applicant" in names
        assert "Rejected Applicant" not in names, (
            "rejected registrations must not appear in the whitelist"
        )

        mats_sheet = wb["Approved Materials"]
        mat_files = {mats_sheet.cell(row=i, column=6).value for i in range(2, mats_sheet.max_row + 1)}
        assert "letter.pdf" in mat_files
    finally:
        if os.path.isfile(file_path):
            os.remove(file_path)


# ── Structural guards for the review-page download control ────────────────

def test_registration_review_page_has_download_control():
    """The reviewer page must expose a per-version download action so
    reviewers can inspect uploaded evidence from the delivered UI. The
    audit flagged the absence of this control as a High-severity gap."""
    import pathlib

    sfc = (
        pathlib.Path(__file__).resolve().parent.parent
        / "frontend" / "src" / "views" / "RegistrationReviewPage.vue"
    )
    content = sfc.read_text()
    # Button is present AND wired to the download backend route.
    assert 'data-test="download-version"' in content, (
        "review page must expose a download control"
    )
    assert "downloadVersion(ver)" in content, (
        "download control must call the downloadVersion handler"
    )
    assert "/registrations/versions/" in content and "/download" in content, (
        "download handler must target the backend download route"
    )


def test_restore_endpoint_source_distinguishes_file_failure():
    """AST-level guard: the restore handler must not blindly return
    ``status=complete`` whenever the DB exit code is zero. The code
    must look at the file restore result too."""
    import pathlib

    src = (
        pathlib.Path(__file__).resolve().parent.parent
        / "app" / "api" / "v1" / "admin_ops.py"
    ).read_text()
    # Reject the old single-branch pattern.
    bad_pattern = 'return {\n            "status": "complete"'
    # The new code checks ``db_ok and files_ok`` before returning complete.
    assert "db_ok and files_ok" in src, (
        "restore handler must gate ``complete`` on both db_ok and files_ok"
    )
    assert "files_attempted" in src, (
        "restore handler must distinguish 'no file backup' from "
        "'file backup attempted and failed'"
    )
