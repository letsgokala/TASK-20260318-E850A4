"""Excel report generation using openpyxl."""
import os
import uuid
from datetime import datetime, timezone
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.audit_log import AuditLog
from app.models.checklist_item import ChecklistItem
from app.models.financial import FinancialTransaction, FundingAccount, TransactionType
from app.models.material import Material, MaterialVersion, MaterialVersionStatus
from app.models.registration import Registration, RegistrationStatus
from app.models.review_record import ReviewRecord

_EXPORT_ROOT = "/storage/exports"
_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def _write_header(ws, headers: list[str]) -> None:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL


async def generate_reconciliation_report(
    db: AsyncSession,
    batch_id: uuid.UUID | None,
    from_date: datetime | None,
    to_date: datetime | None,
) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation"

    headers = [
        "Registration ID", "Applicant Name", "Status",
        "Funding Account", "Allocated Budget",
        "Total Income", "Total Expenses", "Balance", "Overspending",
    ]
    _write_header(ws, headers)

    query = (
        select(
            Registration.id,
            Registration.applicant_name,
            Registration.status,
            FundingAccount.name,
            FundingAccount.allocated_budget,
            FundingAccount.id.label("account_id"),
        )
        .outerjoin(FundingAccount, FundingAccount.registration_id == Registration.id)
        .where(Registration.status != RegistrationStatus.DRAFT)
    )
    if batch_id:
        query = query.where(Registration.batch_id == batch_id)
    if from_date:
        query = query.where(Registration.created_at >= from_date)
    if to_date:
        query = query.where(Registration.created_at <= to_date)

    result = await db.execute(query)
    rows = result.all()

    for row_idx, row in enumerate(rows, 2):
        income = Decimal("0")
        expenses = Decimal("0")
        if row.account_id:
            inc_res = await db.execute(
                select(func.coalesce(func.sum(FinancialTransaction.amount), Decimal("0")))
                .where(
                    FinancialTransaction.funding_account_id == row.account_id,
                    FinancialTransaction.type == TransactionType.INCOME,
                )
            )
            income = inc_res.scalar_one()
            exp_res = await db.execute(
                select(func.coalesce(func.sum(FinancialTransaction.amount), Decimal("0")))
                .where(
                    FinancialTransaction.funding_account_id == row.account_id,
                    FinancialTransaction.type == TransactionType.EXPENSE,
                )
            )
            expenses = exp_res.scalar_one()

        budget = row.allocated_budget or Decimal("0")
        balance = budget + income - expenses
        overspending = "Yes" if expenses > budget + income else "No"

        ws.cell(row=row_idx, column=1, value=str(row.id))
        ws.cell(row=row_idx, column=2, value=row.applicant_name or "")
        ws.cell(row=row_idx, column=3, value=row.status.value if row.status else "")
        ws.cell(row=row_idx, column=4, value=row.name or "N/A")
        ws.cell(row=row_idx, column=5, value=float(budget))
        ws.cell(row=row_idx, column=6, value=float(income))
        ws.cell(row=row_idx, column=7, value=float(expenses))
        ws.cell(row=row_idx, column=8, value=float(balance))
        ws.cell(row=row_idx, column=9, value=overspending)

    return _save_workbook(wb, "reconciliation")


async def generate_audit_report(
    db: AsyncSession,
    from_date: datetime | None,
    to_date: datetime | None,
) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Log"

    headers = [
        "Timestamp", "User ID", "Action",
        "Resource Type", "Resource ID", "Details", "IP Address",
    ]
    _write_header(ws, headers)

    query = select(AuditLog).order_by(AuditLog.created_at.desc())
    if from_date:
        query = query.where(AuditLog.created_at >= from_date)
    if to_date:
        query = query.where(AuditLog.created_at <= to_date)

    result = await db.execute(query)
    logs = result.scalars().all()

    for row_idx, log in enumerate(logs, 2):
        ws.cell(row=row_idx, column=1, value=str(log.created_at))
        ws.cell(row=row_idx, column=2, value=str(log.user_id) if log.user_id else "")
        ws.cell(row=row_idx, column=3, value=log.action)
        ws.cell(row=row_idx, column=4, value=log.resource_type or "")
        ws.cell(row=row_idx, column=5, value=str(log.resource_id) if log.resource_id else "")
        ws.cell(row=row_idx, column=6, value=str(log.details) if log.details else "")
        ws.cell(row=row_idx, column=7, value=str(log.ip_address) if log.ip_address else "")

    return _save_workbook(wb, "audit")


async def generate_compliance_report(
    db: AsyncSession,
    batch_id: uuid.UUID | None,
    from_date: datetime | None,
    to_date: datetime | None,
) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Compliance"

    headers = [
        "Registration ID", "Applicant Name", "Status",
        "Materials Complete", "Review History Summary", "Flagged Issues",
    ]
    _write_header(ws, headers)

    query = select(Registration).where(Registration.status != RegistrationStatus.DRAFT)
    if batch_id:
        query = query.where(Registration.batch_id == batch_id)
    if from_date:
        query = query.where(Registration.created_at >= from_date)
    if to_date:
        query = query.where(Registration.created_at <= to_date)

    result = await db.execute(query)
    regs = result.scalars().all()

    for row_idx, reg in enumerate(regs, 2):
        # Check material completeness against the batch checklist.
        # A registration is "Materials Complete" only when every
        # REQUIRED checklist item has at least one uploaded version
        # whose status is not needs_correction. Optional items are
        # not required for completeness. The prior count-based check
        # was too lax and could report "Yes" when required items were
        # missing.
        required_items_result = await db.execute(
            select(ChecklistItem).where(
                ChecklistItem.batch_id == reg.batch_id,
                ChecklistItem.is_required == True,
            )
        )
        required_items = required_items_result.scalars().all()

        materials_complete = "Yes"
        if not required_items:
            # No required items → treat as complete (nothing is demanded).
            pass
        else:
            for ci in required_items:
                mat_result = await db.execute(
                    select(Material).where(
                        Material.registration_id == reg.id,
                        Material.checklist_item_id == ci.id,
                    )
                )
                mat = mat_result.scalar_one_or_none()
                if not mat:
                    materials_complete = "No"
                    break
                # At least one version that is not needs_correction.
                valid_ver = await db.execute(
                    select(MaterialVersion.id).where(
                        MaterialVersion.material_id == mat.id,
                        MaterialVersion.status != MaterialVersionStatus.NEEDS_CORRECTION,
                    ).limit(1)
                )
                if not valid_ver.scalar_one_or_none():
                    materials_complete = "No"
                    break

        # Review history summary
        review_result = await db.execute(
            select(ReviewRecord)
            .where(ReviewRecord.registration_id == reg.id)
            .order_by(ReviewRecord.reviewed_at)
        )
        reviews = review_result.scalars().all()
        history_summary = " -> ".join(
            f"{r.from_status}->{r.to_status}" for r in reviews
        ) if reviews else "No reviews"

        # Flagged issues
        issues = []
        if reg.supplementary_used:
            issues.append("Supplementary submission used")
        dup_result = await db.execute(
            select(func.count()).select_from(MaterialVersion)
            .join(Material, MaterialVersion.material_id == Material.id)
            .where(Material.registration_id == reg.id, MaterialVersion.duplicate_flag == True)
        )
        dup_count = dup_result.scalar_one()
        if dup_count > 0:
            issues.append(f"{dup_count} duplicate file(s)")

        ws.cell(row=row_idx, column=1, value=str(reg.id))
        ws.cell(row=row_idx, column=2, value=reg.applicant_name or "")
        ws.cell(row=row_idx, column=3, value=reg.status.value)
        ws.cell(row=row_idx, column=4, value=materials_complete)
        ws.cell(row=row_idx, column=5, value=history_summary)
        ws.cell(row=row_idx, column=6, value="; ".join(issues) if issues else "None")

    return _save_workbook(wb, "compliance")


async def generate_whitelist_report(
    db: AsyncSession,
    batch_id: uuid.UUID | None,
) -> str:
    """Whitelist report — approved registrations and their materials per batch.

    The design doc and the prompt both frame the whitelist as "the approved
    set" — registrations that cleared review, paired with the materials they
    submitted as evidence. The previous implementation exported the batch
    CHECKLIST definitions only, which drifted from the documented intent
    (the audit report flagged this as a High-visibility semantics gap).

    Output is a two-sheet workbook:

    1. "Approved Registrations" — one row per approved registration with
       batch, applicant, status, and approval timestamp.
    2. "Approved Materials" — one row per material version of an approved
       registration, so auditors can reconcile the documentary evidence
       against the approval list.
    """
    from app.models.collection_batch import CollectionBatch
    from app.models.checklist_item import ChecklistItem

    wb = Workbook()

    # ── Sheet 1: Approved Registrations ────────────────────────────────
    ws = wb.active
    ws.title = "Approved Registrations"
    _write_header(ws, [
        "Batch ID", "Batch Name",
        "Registration ID", "Applicant Name", "Title",
        "Status", "Approved At",
    ])

    # Pick approved + promoted_from_waitlist — both represent "on the
    # whitelist for funding" per the registration state machine.
    approved_statuses = [
        RegistrationStatus.APPROVED,
        RegistrationStatus.PROMOTED_FROM_WAITLIST,
    ]
    reg_query = (
        select(Registration, CollectionBatch)
        .join(CollectionBatch, Registration.batch_id == CollectionBatch.id)
        .where(Registration.status.in_(approved_statuses))
    )
    if batch_id:
        reg_query = reg_query.where(Registration.batch_id == batch_id)
    reg_query = reg_query.order_by(
        CollectionBatch.created_at.desc(),
        Registration.updated_at.desc(),
    )
    reg_rows = (await db.execute(reg_query)).all()

    approved_ids: list[uuid.UUID] = []
    row_idx = 2
    for reg, batch in reg_rows:
        approved_ids.append(reg.id)
        # Use the latest approval-style review record (if any) for the
        # "approved at" column; fall back to the registration's
        # updated_at.
        review_result = await db.execute(
            select(ReviewRecord)
            .where(
                ReviewRecord.registration_id == reg.id,
                ReviewRecord.to_status.in_([s.value for s in approved_statuses]),
            )
            .order_by(ReviewRecord.reviewed_at.desc())
            .limit(1)
        )
        rec = review_result.scalar_one_or_none()
        approved_at = rec.reviewed_at if rec else reg.updated_at

        ws.cell(row=row_idx, column=1, value=str(batch.id))
        ws.cell(row=row_idx, column=2, value=batch.name)
        ws.cell(row=row_idx, column=3, value=str(reg.id))
        ws.cell(row=row_idx, column=4, value=reg.applicant_name or "")
        ws.cell(row=row_idx, column=5, value=reg.title or "")
        ws.cell(row=row_idx, column=6, value=reg.status.value)
        ws.cell(row=row_idx, column=7, value=str(approved_at) if approved_at else "")
        row_idx += 1

    # ── Sheet 2: Approved Materials ────────────────────────────────────
    mat_ws = wb.create_sheet("Approved Materials")
    _write_header(mat_ws, [
        "Batch ID", "Registration ID", "Applicant Name",
        "Checklist Item", "Version", "Filename", "Status",
        "SHA-256", "Size (bytes)", "Uploaded At",
    ])

    mat_row = 2
    if approved_ids:
        # One multi-row join is cheaper than N per-reg lookups, and avoids
        # drifting out of sync with the approved list above.
        mat_query = (
            select(
                Registration.batch_id,
                Registration.id.label("reg_id"),
                Registration.applicant_name,
                ChecklistItem.label.label("checklist_label"),
                MaterialVersion.version_number,
                MaterialVersion.original_filename,
                MaterialVersion.status.label("version_status"),
                MaterialVersion.sha256_hash,
                MaterialVersion.file_size_bytes,
                MaterialVersion.uploaded_at,
            )
            .join(Material, Material.registration_id == Registration.id)
            .join(ChecklistItem, ChecklistItem.id == Material.checklist_item_id)
            .join(MaterialVersion, MaterialVersion.material_id == Material.id)
            .where(Registration.id.in_(approved_ids))
            .order_by(
                Registration.id,
                ChecklistItem.sort_order,
                MaterialVersion.version_number,
            )
        )
        for row in (await db.execute(mat_query)).all():
            mat_ws.cell(row=mat_row, column=1, value=str(row.batch_id))
            mat_ws.cell(row=mat_row, column=2, value=str(row.reg_id))
            mat_ws.cell(row=mat_row, column=3, value=row.applicant_name or "")
            mat_ws.cell(row=mat_row, column=4, value=row.checklist_label)
            mat_ws.cell(row=mat_row, column=5, value=row.version_number)
            mat_ws.cell(row=mat_row, column=6, value=row.original_filename)
            mat_ws.cell(
                row=mat_row, column=7,
                value=row.version_status.value if hasattr(row.version_status, "value") else str(row.version_status),
            )
            mat_ws.cell(row=mat_row, column=8, value=row.sha256_hash)
            mat_ws.cell(row=mat_row, column=9, value=row.file_size_bytes)
            mat_ws.cell(row=mat_row, column=10, value=str(row.uploaded_at) if row.uploaded_at else "")
            mat_row += 1

    return _save_workbook(wb, "whitelist")


def _save_workbook(wb: Workbook, prefix: str) -> str:
    os.makedirs(_EXPORT_ROOT, exist_ok=True)
    task_id = str(uuid.uuid4())
    file_path = os.path.join(_EXPORT_ROOT, f"{task_id}.xlsx")
    wb.save(file_path)
    return file_path
